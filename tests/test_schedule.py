"""Date-distribution feature: capacity spread, classification, distribution, export."""
import datetime
import io

from openpyxl import Workbook, load_workbook

from app import scheduling as S


def test_day_capacities_spread_evenly():
    caps = S.day_capacities(163, 21)
    assert sum(caps) == 163
    assert max(caps) - min(caps) <= 1          # base or base+1 only
    caps2 = S.day_capacities(100, 30)
    assert sum(caps2) == 100 and max(caps2) - min(caps2) <= 1
    # fewer links than days -> spread out, no day > 1
    caps3 = S.day_capacities(3, 21)
    assert sum(caps3) == 3 and max(caps3) == 1


def test_classify():
    # naked URL -> always anchorless filler; bare domain / brand -> branded
    assert S.classify("https://x.com/", "ND") == S.ANCHORLESS
    assert S.classify("x.com", "") == S.BRANDED          # bare domain = BD -> branded
    assert S.classify("BrandName", "BD") == S.BRANDED
    assert S.classify("online casino österreich", "EM") == S.COMMERCIAL
    assert S.classify("casino per magenta", "PM") == S.COMMERCIAL
    assert S.classify("Casino Vergleich von Brand", "BD+PM") == S.COMMERCIAL


def _plan_book(rows):
    wb = Workbook(); ws = wb.active
    ws.append(["Sprint", "SEO Specialist", "Project", "Project Url", "URL Type",
               "Link Type", "Anchor Type", "Anchor", "Keyword"])
    for r in rows:
        ws.append(r)
    b = io.BytesIO(); wb.save(b); return b.getvalue()


def test_read_plan_and_distribute():
    rows = []
    for _ in range(60):
        rows.append(["1 S", "Miles", "x.com", "https://x.com/", "Main Page", "BH", "BD",
                     "https://x.com/", "x"])          # anchorless
    for _ in range(9):
        rows.append(["1 S", "Miles", "x.com", "https://x.com/", "Main Page", "BH", "EM",
                     "x casino", "x casino"])          # commercial
    header, links = S.read_plan(_plan_book(rows))
    assert "Sprint" in header and len(links) == 69
    s = S.summarize(links, 30)
    assert s["anchorless"] == 60 and s["commercial"] == 9
    placements = S.distribute(links, 30, datetime.date(2026, 7, 9))
    assert len(placements) == 69
    # every date within [start, start+29]
    days = {(d - datetime.date(2026, 7, 9)).days for d, _ in placements}
    assert min(days) == 0 and max(days) <= 29
    # 9 commercial over 30 days -> spread, never two on the same day
    from collections import Counter
    comm = Counter(d for d, l in placements if l["category"] == S.COMMERCIAL)
    assert max(comm.values()) == 1 and len(comm) == 9


def test_build_scheduled_workbook_replaces_sprint_with_date():
    rows = [["1 S", "Miles", "x.com", "https://x.com/", "Main Page", "BH", "BD",
             "https://x.com/", "x"] for _ in range(10)]
    header, links = S.read_plan(_plan_book(rows))
    placements = S.distribute(links, 5, datetime.date(2026, 7, 9))
    content = S.build_scheduled_workbook(header, placements)
    ws = load_workbook(io.BytesIO(content)).active
    out_header = [c.value for c in ws[1]]
    assert "Date" in out_header and "Sprint" not in out_header
    date_idx = out_header.index("Date")
    assert [c.value for c in ws[2]][date_idx] == "09.07.2026"
