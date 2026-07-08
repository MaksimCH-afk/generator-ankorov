"""End-to-end endpoint tests via FastAPI TestClient (isolated temp DB)."""
from fastapi.testclient import TestClient

from app.database import SessionLocal
from app.main import app
from app.models import Project, Strategy


def _project_id(url: str) -> int:
    db = SessionLocal()
    try:
        return db.query(Project).filter(Project.url == url).first().id
    finally:
        db.close()


def test_health_and_pages():
    with TestClient(app) as c:
        assert c.get("/health").json()["status"] == "ok"
        for p in ["/", "/generate", "/schedule", "/anchor-types", "/strategies", "/profiles",
                  "/suffixes", "/history", "/logs", "/anchors"]:
            assert c.get(p).status_code == 200


def test_create_dedup_and_normalize():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "http://www.dedup-test.com", "language": "English"},
               follow_redirects=False)
        # www stripped + https forced + trailing slash
        pid = _project_id("https://dedup-test.com/")
        assert pid
        # duplicate (different form, same canonical) is rejected
        r = c.post("/projects/create", data={"url": "https://www.dedup-test.com/"}, follow_redirects=False)
        assert "error=" in r.headers["location"]


def test_full_generation_flow_and_download():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://flow.com/", "language": "German"},
               follow_redirects=False)
        pid = _project_id("https://flow.com/")
        # upload keywords
        csv = b"keyword,frequency\nflow,1000\nflow casino,500\nflow login,100\n"
        c.post(f"/projects/{pid}/keywords", files={"file": ("f.csv", csv, "text/csv")},
               follow_redirects=False)
        sid = SessionLocal().query(Strategy).first().id  # "Обычная"
        c.post(f"/projects/{pid}/strategy", data={"strategy_id": sid, "next": f"/projects/{pid}"},
               follow_redirects=False)
        c.post(f"/projects/{pid}/volume", data={"volume": 50}, follow_redirects=False)
        # generate -> JSON token
        r = c.post("/generate", data={"project_ids": [pid], "sprint": "122",
                                      "export_format": "separate", "group_mode": "expand"})
        assert r.status_code == 200
        body = r.json()
        assert body["total"] == 50 and body["count"] == 1
        token = body["token"]
        # download once -> xlsx; twice -> gone
        d = c.get(f"/generate/download/{token}")
        assert d.status_code == 200 and "spreadsheet" in d.headers["content-type"]
        d2 = c.get(f"/generate/download/{token}", follow_redirects=False)
        assert d2.status_code == 303


def test_export_and_breakdown():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://exp.com/", "language": ""}, follow_redirects=False)
        pid = _project_id("https://exp.com/")
        c.post(f"/projects/{pid}/keywords",
               files={"file": ("f.csv", b"keyword,frequency\nexp,100\n", "text/csv")}, follow_redirects=False)
        sid = SessionLocal().query(Strategy).first().id
        c.post(f"/projects/{pid}/strategy", data={"strategy_id": sid}, follow_redirects=False)
        c.post(f"/projects/{pid}/volume", data={"volume": 20}, follow_redirects=False)
        # breakdown JSON
        rows = c.get(f"/projects/{pid}/breakdown").json()["rows"]
        assert rows and sum(r["count"] for r in rows) == 20
        # one-click export
        e = c.get(f"/projects/{pid}/export")
        assert e.status_code == 200 and "spreadsheet" in e.headers["content-type"]


def test_delete_removes_from_generation():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://del.com/", "brand": "del"}, follow_redirects=False)
        pid = _project_id("https://del.com/")
        assert "https://del.com/" in c.get("/generate").text
        c.post(f"/projects/{pid}/delete", follow_redirects=False)
        assert "https://del.com/" not in c.get("/generate").text     # gone from generation too
        assert SessionLocal().get(Project, pid) is None


def test_reclassify_all_updates_stored_recognition():
    from app.models import Keyword
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://ract.com/", "brand": "ract"}, follow_redirects=False)
        pid = _project_id("https://ract.com/")
        # insert keywords directly WITHOUT recognition (simulates a pre-feature project)
        db = SessionLocal()
        for i, kw in enumerate(["ract casino", "ract no deposit bonus"]):
            db.add(Keyword(project_id=pid, keyword=kw, frequency=100 - i, position=i))
        db.commit(); db.close()
        # nothing recognised yet
        db = SessionLocal()
        assert all(not k.excluded and not k.anchor_type
                   for k in db.query(Keyword).filter_by(project_id=pid).all())
        db.close()
        c.post("/projects/reclassify-all", follow_redirects=False)
        db = SessionLocal()
        kws = {k.keyword: k for k in db.query(Keyword).filter_by(project_id=pid).all()}
        assert kws["ract no deposit bonus"].excluded is True   # stop-anchor now flagged
        assert all(k.anchor_type for k in kws.values())         # types now filled
        db.close()


def test_settings_key_and_per_action_models():
    from app import appsettings
    from app.database import SessionLocal
    with TestClient(app) as c:
        # recommended defaults are returned before anything is saved
        db = SessionLocal()
        assert appsettings.get_action_model(db, "types") == appsettings.RECOMMENDED["types"]
        db.close()
        # save a key + custom models per action
        data = {"key": "sk-or-v1-testkey1234", "model_jokes": "openai/gpt-4o-mini",
                "model_types": "anthropic/claude-3.5-sonnet", "model_smart": "openai/gpt-4o-mini",
                "model_schedule_smart": "openai/gpt-4o", "model_schedule_cheap": "openai/gpt-4o-mini"}
        r = c.post("/settings/save", data=data, follow_redirects=False)
        assert r.status_code == 303
        db = SessionLocal()
        assert appsettings.get_key(db) == "sk-or-v1-testkey1234"
        assert appsettings.get_action_model(db, "types") == "anthropic/claude-3.5-sonnet"
        assert appsettings.get_action_slot(db, "jokes") == ("sk-or-v1-testkey1234", "openai/gpt-4o-mini")
        db.close()
        # dashboard shows the settings block + saved-key badge
        assert "Проставить рекомендуемые" in c.get("/").text
        # clearing the key
        c.post("/settings/save", data={"action": "clear_key"}, follow_redirects=False)
        db = SessionLocal()
        assert appsettings.get_key(db) == ""
        db.close()


def test_upload_classifies_and_generation_excludes_stopanchors():
    from app.models import Keyword
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://stp.com/", "brand": "stp"}, follow_redirects=False)
        pid = _project_id("https://stp.com/")
        # "no deposit bonus" is a default stop-anchor -> should be excluded at upload
        csv = b"keyword,frequency\nstp casino,1000\nstp,500\nstp no deposit bonus,200\n"
        c.post(f"/projects/{pid}/keywords", files={"file": ("f.csv", csv, "text/csv")}, follow_redirects=False)
        db = SessionLocal()
        kws = {k.keyword: k for k in db.query(Keyword).filter_by(project_id=pid).all()}
        assert kws["stp no deposit bonus"].excluded is True     # stop-anchor flagged at upload
        assert kws["stp casino"].excluded is False
        assert all(k.anchor_type for k in kws.values())          # types stored
        db.close()
        sid = SessionLocal().query(Strategy).first().id
        c.post(f"/projects/{pid}/strategy", data={"strategy_id": sid}, follow_redirects=False)
        c.post(f"/projects/{pid}/volume", data={"volume": 100}, follow_redirects=False)
        # export params live in Settings now: enable stop-anchor exclusion + grouping
        c.post("/settings/save", data={"gen_smart": "on", "gen_group": "group"}, follow_redirects=False)
        # generation drops the excluded keyword; runs offline
        r = c.post("/generate", data={"project_ids": [pid]})
        assert r.status_code == 200
        d = c.get(f"/generate/download/{r.json()['token']}")
        anchors = _anchor_column(d.content)
        assert "stp no deposit bonus" not in anchors and "stp casino" in anchors


def _anchor_column(xlsx_bytes):
    import io as _io
    from openpyxl import load_workbook
    from app.excel_export import BASE_COLUMNS
    ws = load_workbook(_io.BytesIO(xlsx_bytes)).active
    # grouped export -> leading Link Q-ty column
    off = 1 if [c.value for c in ws[1]][0] == "Link Q-ty" else 0
    ai = BASE_COLUMNS.index("Anchor") + off
    return {[c.value for c in ws[r]][ai] for r in range(2, ws.max_row + 1)}


def test_duplicate_project_reuses_keywords():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://dup-src.com/", "language": "German",
                                         "brand": "DupBrand"}, follow_redirects=False)
        pid = _project_id("https://dup-src.com/")
        c.post(f"/projects/{pid}/keywords",
               files={"file": ("f.csv", b"keyword,frequency\na,100\nb,50\nc,10\n", "text/csv")},
               follow_redirects=False)
        sid = SessionLocal().query(Strategy).first().id
        c.post(f"/projects/{pid}/strategy", data={"strategy_id": sid}, follow_redirects=False)
        c.post(f"/projects/{pid}/volume", data={"volume": 250}, follow_redirects=False)
        # duplicate to two mirror domains, keep the source strategy
        r = c.post(f"/projects/{pid}/duplicate",
                   data={"domains": "https://dup-m1.com/\ndup-m2.at", "strategy": "keep"},
                   follow_redirects=False)
        assert r.status_code == 303
        db = SessionLocal()
        for url in ("https://dup-m1.com/", "https://dup-m2.at/"):
            clone = db.query(Project).filter(Project.url == url).first()
            assert clone is not None
            assert len(clone.keywords) == 3          # keywords copied
            assert clone.language == "German" and clone.brand == "DupBrand"
            assert clone.volume == 250 and clone.strategy_id == sid
        db.close()
        # duplicating onto an existing domain is skipped
        r2 = c.post(f"/projects/{pid}/duplicate",
                    data={"domains": "https://dup-m1.com/", "strategy": "keep"}, follow_redirects=False)
        assert "location" in r2.headers


def test_schedule_multiple_files_returns_zip():
    import io as _io
    import zipfile
    from openpyxl import Workbook

    def plan_book(project):
        wb = Workbook(); ws = wb.active
        ws.append(["Sprint", "SEO Specialist", "Project", "Project Url", "URL Type",
                   "Link Type", "Anchor Type", "Anchor", "Keyword"])
        for _ in range(20):
            ws.append(["1 S", "M", project, f"https://{project}/", "Main Page", "BH",
                       "ND", f"https://{project}/", ""])
        b = _io.BytesIO(); wb.save(b); return b.getvalue()

    with TestClient(app) as c:
        r = c.post("/schedule/generate",
                   data={"days": "10", "start_date": "2026-07-09", "use_model": ""},
                   files=[("files", ("a.xlsx", plan_book("a-casino.com"), "application/octet-stream")),
                          ("files", ("b.xlsx", plan_book("b-casino.com"), "application/octet-stream"))],
                   follow_redirects=False)
        assert r.status_code == 200 and "zip" in r.headers["content-type"]
        zf = zipfile.ZipFile(_io.BytesIO(r.content))
        assert len(zf.namelist()) == 2  # one scheduled file per input

        # per-file periods/start dates align to upload order
        r2 = c.post("/schedule/generate",
                    data={"days": "30", "start_date": "2026-07-09", "use_model": "",
                          "per_file": "on", "per_days": ["10", "20"],
                          "per_start": ["2026-07-01", "2026-08-01"]},
                    files=[("files", ("a.xlsx", plan_book("a-casino.com"), "application/octet-stream")),
                           ("files", ("b.xlsx", plan_book("b-casino.com"), "application/octet-stream"))],
                    follow_redirects=False)
        names = zipfile.ZipFile(_io.BytesIO(r2.content)).namelist()
        assert "a-2026-07-01-10d.xlsx" in names   # file a: 10 days from Jul 1
        assert "b-2026-08-01-20d.xlsx" in names   # file b: 20 days from Aug 1


def test_schedule_history_saved_and_downloadable():
    import io as _io
    from openpyxl import Workbook
    from app.models import ScheduleRun

    def plan_book():
        wb = Workbook(); ws = wb.active
        ws.append(["Sprint", "SEO Specialist", "Project", "Project Url", "URL Type",
                   "Link Type", "Anchor Type", "Anchor", "Keyword"])
        for _ in range(12):
            ws.append(["1 S", "M", "h.com", "https://h.com/", "Main Page", "BH", "ND", "https://h.com/", ""])
        b = _io.BytesIO(); wb.save(b); return b.getvalue()

    with TestClient(app) as c:
        r = c.post("/schedule/generate",
                   data={"days": "6", "start_date": "2026-07-09", "use_model": ""},
                   files=[("files", ("hist.xlsx", plan_book(), "application/octet-stream"))],
                   follow_redirects=False)
        assert r.status_code == 200
        produced = r.content
        db = SessionLocal()
        run = db.query(ScheduleRun).order_by(ScheduleRun.id.desc()).first()
        assert run is not None and run.content == produced   # exact result stored
        rid = run.id
        db.close()
        # re-download later returns the same bytes
        d = c.get(f"/schedule/history/{rid}")
        assert d.status_code == 200 and d.content == produced
        # delete removes it
        c.post(f"/schedule/history/{rid}/delete", follow_redirects=False)
        assert SessionLocal().get(ScheduleRun, rid) is None


def test_bulk_and_anchors():
    with TestClient(app) as c:
        c.post("/projects/create", data={"url": "https://b1.com/"}, follow_redirects=False)
        c.post("/projects/create", data={"url": "https://b2.com/"}, follow_redirects=False)
        id1, id2 = _project_id("https://b1.com/"), _project_id("https://b2.com/")
        r = c.post("/projects/bulk-delete", data={"project_ids": [id1, id2]}, follow_redirects=False)
        assert r.status_code == 303
        db = SessionLocal()
        assert db.get(Project, id1) is None and db.get(Project, id2) is None
        db.close()
        # anchors add + delete
        c.post("/anchors/add", data={"phrases": "casino spam\nwelcome bonus"}, follow_redirects=False)
        assert c.get("/anchors").status_code == 200
