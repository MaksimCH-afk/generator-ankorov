"""Tests for the Excel export (columns, expansion, grouping, language, url type)."""
import io

from openpyxl import load_workbook

from app.anchortypes import classify_one
from app.excel_export import BASE_COLUMNS, build_workbook
from app.generator import GeneratedRow


def _rows():
    return {
        "Прогоны": [
            GeneratedRow(link_qty=3, url="https://betalice.com/", anchor="https://betalice.com/",
                         article_language="German", keyword="https://betalice.com/"),
            GeneratedRow(link_qty=2, url="https://betalice.com/", anchor="betalice",
                         article_language="German", keyword="betalice", is_keyword=True),
        ],
        "Внутренние страницы": [
            GeneratedRow(link_qty=1, url="https://betalice.com/boni/", anchor="betalice auszahlung",
                         article_language="German", keyword="betalice auszahlung"),
        ],
    }


def _load(content):
    return load_workbook(io.BytesIO(content))


def test_expanded_columns_and_row_count():
    wb = _load(build_workbook(_rows(), sprint="122", seo_specialist="Miles Nashwood", language="German"))
    ws = wb["Прогоны"]
    header = [c.value for c in ws[1]]
    assert header == BASE_COLUMNS + ["Article Language"]
    # 3 + 2 links expanded -> 5 data rows
    assert ws.max_row - 1 == 5
    first = [c.value for c in ws[2]]
    assert first[0] == "122"                       # Sprint
    assert first[1] == "Miles Nashwood"            # SEO Specialist
    assert first[2] == "betalice.com"              # Project = bare domain
    assert first[3] == "https://betalice.com/"     # Project Url
    assert first[4] == "Main Page"                 # URL Type
    assert first[-1] == "German"                   # Article Language


def test_grouped_has_link_qty_column():
    wb = _load(build_workbook(_rows(), grouped=True, language="German"))
    ws = wb["Прогоны"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Link Q-ty"
    # grouped: one row per anchor -> 2 rows
    assert ws.max_row - 1 == 2
    assert [c.value for c in ws[2]][0] == 3        # quantity in first column


def test_language_omitted_when_empty():
    wb = _load(build_workbook(_rows(), language=""))
    header = [c.value for c in wb["Прогоны"][1]]
    assert "Article Language" not in header


def test_internal_sheet_url_type_inner_page():
    wb = _load(build_workbook(_rows(), language="German"))
    ws = wb["Внутренние страницы"]
    row = [c.value for c in ws[2]]
    assert row[4] == "Inner Page"
    assert row[3] == "https://betalice.com/boni/"


def test_keyword_column_holds_top_keyword_on_every_row():
    kw_idx = BASE_COLUMNS.index("Keyword")
    wb = _load(build_workbook(_rows(), language="German"))
    # Every row in every sheet carries the project's most-used keyword.
    for sheet in ("Прогоны", "Внутренние страницы"):
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            assert [c.value for c in ws[r]][kw_idx] == "betalice"


def test_anchor_type_7type_classification():
    kws = {"online casino", "klarna casino", "betalice casino"}
    # ND: naked URL with protocol
    assert classify_one("https://betalice.com/") == "ND"
    # BD: bare domain without protocol, or the brand name
    assert classify_one("betalice.com") == "BD"
    assert classify_one("AustriaWin24.at") == "BD"
    assert classify_one("betalice", brand="betalice") == "BD"
    # EM: exact keyword
    assert classify_one("online casino", keywords=kws) == "EM"
    # PM: keyword diluted in a phrase
    assert classify_one("bestes online casino für alle", keywords=kws) == "PM"
    # BD+PM: brand + partial keyword
    assert classify_one("Casino Vergleich von Betalice", brand="Betalice", keywords=kws) == "BD+PM"
    # G: thematic, no keyword
    assert classify_one("geprüfte deutsche Spielportale", keywords=kws) == "G"
    # NT: empty anchor
    assert classify_one("") == "NT"


def test_anchor_type_column_in_export():
    at_idx = BASE_COLUMNS.index("Anchor Type")
    wb = _load(build_workbook(_rows(), language="German", brand="betalice"))
    ws = wb["Прогоны"]
    valid = {"ND", "BD", "EM", "PM", "G", "BD+PM", "NT"}
    for r in range(2, ws.max_row + 1):
        assert [c.value for c in ws[r]][at_idx] in valid
    # the anchorless https:// rows are ND now
    types = {[c.value for c in ws[r]][at_idx] for r in range(2, ws.max_row + 1)}
    assert "ND" in types


def test_language_column_suppressed_by_flag():
    # Even with a language set, include_language=False drops the column.
    wb = _load(build_workbook(_rows(), language="German", include_language=False))
    header = [c.value for c in wb["Прогоны"][1]]
    assert "Article Language" not in header


def test_crowd_keyword_fallback_fills_column():
    kw_idx = BASE_COLUMNS.index("Keyword")
    # Crowd campaign: only anchorless rows (no keyword anchors).
    sheets = {
        "Крауд+сабмиты": [
            GeneratedRow(link_qty=3, url="https://x.com/", anchor="x.com",
                         article_language="", keyword="x.com"),
            GeneratedRow(link_qty=2, url="https://x.com/", anchor="https://x.com/",
                         article_language="", keyword="https://x.com/"),
        ],
    }
    wb = _load(build_workbook(sheets, keyword="best keyword"))
    ws = wb["Крауд+сабмиты"]
    for r in range(2, ws.max_row + 1):
        assert [c.value for c in ws[r]][kw_idx] == "best keyword"


def test_keyword_blank_for_fully_anchorless():
    kw_idx = BASE_COLUMNS.index("Keyword")
    sheets = {
        "Крауд+сабмиты": [
            GeneratedRow(link_qty=2, url="https://x.com/", anchor="https://x.com/",
                         article_language="", keyword="https://x.com/"),
        ],
    }
    wb = _load(build_workbook(sheets, language=""))
    ws = wb["Крауд+сабмиты"]
    assert [c.value for c in ws[2]][kw_idx] in (None, "")
