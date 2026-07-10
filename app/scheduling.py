"""Date distribution ("Распределение по датам").

Independent feature: takes a previously-exported link plan (Excel), spreads its
links evenly across a date range, and re-exports with a ``Date`` column instead
of ``Sprint``.

Distribution rules (see spec):
* total links / days = average per day; the remainder is **not** dumped on one
  day — the extra links are spread evenly across the period (every ~days/rem-th
  day gets +1).
* anchorless links (bare URL/domain) are the safe bulk and fill most of each
  day; commercial/branded anchors are spread evenly so they recur regularly but
  never cluster (no spammy day). This is achieved by interleaving all links by
  fractional rank within their category, then slicing the sequence into days.

Anchor category is taken from the plan's ``Anchor Type`` column (BD/EM/PM) plus a
bare-URL check, and can optionally be refined by an OpenRouter model (cheap by
design — a mini model is plenty for 3-way classification).
"""
from __future__ import annotations

import datetime
import io
import re
from collections import defaultdict

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter

from . import anchortypes
from .excel_export import HEADER_FILL, HEADER_FONT

# Cadence buckets used by the distribution: anchorless = safe bulk, branded =
# regular, commercial = spread pointwise.
ANCHORLESS, BRANDED, COMMERCIAL = "anchorless", "branded", "commercial"
_PRIORITY = {COMMERCIAL: 0, BRANDED: 1, ANCHORLESS: 2}

# Company 7-type -> cadence bucket.
_BUCKET = {
    anchortypes.ND: ANCHORLESS, anchortypes.NT: ANCHORLESS,
    anchortypes.BD: BRANDED, anchortypes.G: BRANDED,
    anchortypes.EM: COMMERCIAL, anchortypes.PM: COMMERCIAL, anchortypes.BDPM: COMMERCIAL,
}


def bucket_of(type_code: str) -> str:
    return _BUCKET.get((type_code or "").strip().upper(), COMMERCIAL)


def classify(anchor: str, anchor_type: str) -> str:
    """Cadence bucket from the plan's Anchor Type column (any of the 7 codes),
    falling back to the deterministic company classifier on the anchor text."""
    if anchortypes.looks_naked_url(anchor):
        return ANCHORLESS  # a naked URL is always safe filler, regardless of label
    at = (anchor_type or "").strip().upper()
    if at in _BUCKET:
        return bucket_of(at)
    return bucket_of(anchortypes.classify_one(anchor))


def _sheet_links(rows: list[list[str]]):
    """Parse one sheet's rows -> ``(header, links, project_idx)`` or ``None``.

    Each link is ``{"row", "anchor", "category", "project"}``. ``Link Q-ty`` is
    expanded to one link per unit; trailing empty header columns are trimmed.
    """
    rows = [r for r in rows if any(r)]
    if not rows:
        return None
    full = rows[0]
    # Trim leading and trailing empty (unlabeled) header columns.
    start = 0
    while start < len(full) and not full[start]:
        start += 1
    end = len(full)
    while end > start and not full[end - 1]:
        end -= 1
    header = full[start:end]
    ncol = len(header)
    idx = {name: i for i, name in enumerate(header)}
    qty_i, anchor_i, at_i, proj_i = idx.get("Link Q-ty"), idx.get("Anchor"), idx.get("Anchor Type"), idx.get("Project")
    links: list[dict] = []
    for raw in rows[1:]:
        row = [(raw[start + i] if start + i < len(raw) else "") for i in range(ncol)]
        anchor = row[anchor_i] if anchor_i is not None else ""
        if not anchor:
            continue
        cat = classify(anchor, row[at_i] if at_i is not None else "")
        n = 1
        if qty_i is not None:
            try:
                n = max(1, int(float(row[qty_i])))
            except (ValueError, TypeError):
                n = 1
        proj = row[proj_i] if proj_i is not None else ""
        for _ in range(n):
            links.append({"row": row[:], "anchor": anchor, "category": cat, "project": proj})
    return header, links, proj_i


def read_plan(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse the active sheet of an exported plan into ``(header, links)``."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else str(c).strip() for c in r] for r in ws.iter_rows(values_only=True)]
    res = _sheet_links(rows)
    if not res:
        return [], []
    header, links, _ = res
    return header, links


def read_project_plans(content: bytes, default_name: str = "plan") -> list[dict]:
    """Parse an exported plan into one sub-plan **per project** across ALL sheets.

    A single workbook often holds many projects (``Project`` column) on one or
    more sheets (names vary). Each distinct project becomes its own plan so it
    can be date-distributed independently. Returns a list of
    ``{"name", "header", "links"}``. Sheets without a ``Project`` column yield a
    single plan named after the sheet.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    plans: list[dict] = []
    for ws in wb.worksheets:
        rows = [["" if c is None else str(c).strip() for c in r] for r in ws.iter_rows(values_only=True)]
        res = _sheet_links(rows)
        if not res:
            continue
        header, links, proj_i = res
        if not links:
            continue
        if proj_i is not None:
            groups: dict[str, list[dict]] = defaultdict(list)
            for l in links:
                groups[l["project"] or (ws.title or default_name)].append(l)
            for name, ls in groups.items():
                plans.append({"name": name, "header": header, "links": ls})
        else:
            plans.append({"name": ws.title or default_name, "header": header, "links": links})
    return plans


def safe_name(name: str) -> str:
    """Filesystem-safe base name for a project/sheet."""
    n = re.sub(r"^https?://", "", (name or "").strip()).strip("/")
    n = re.sub(r"[^A-Za-z0-9._-]+", "_", n)
    return n or "plan"


def day_capacities(total: int, days: int) -> list[int]:
    """How many links each day carries: base = total//days, remainder spread
    evenly (every ~days/rem-th day gets +1). Sums to ``total``."""
    if days <= 0:
        return []
    base, rem = divmod(total, days)
    caps = [base] * days
    for k in range(rem):
        d = min(int((k + 0.5) * days / rem), days - 1)
        caps[d] += 1
    return caps


def order_links(links: list[dict]) -> list[dict]:
    """Interleave links so that **each distinct anchor** is spread evenly across
    the whole sequence (fractional-rank merge per anchor text), not grouped one
    anchor after another. Category priority only breaks ties at equal rank."""
    groups: dict[str, list[dict]] = defaultdict(list)
    for l in links:
        groups[l["anchor"]].append(l)  # group by distinct anchor text
    ranked = []
    for _anchor, items in groups.items():
        n = len(items)
        prio = _PRIORITY.get(items[0]["category"], 3)
        for i, it in enumerate(items):
            ranked.append(((i + 0.5) / n, prio, it))
    ranked.sort(key=lambda x: (x[0], x[1]))
    return [it for _, _, it in ranked]


def distribute(links: list[dict], days: int, start_date: datetime.date) -> list[tuple[datetime.date, dict]]:
    """Assign each link a date. Returns ``[(date, link), ...]`` in calendar order."""
    caps = day_capacities(len(links), days)
    ordered = order_links(links)
    out: list[tuple[datetime.date, dict]] = []
    pos = 0
    for d, cap in enumerate(caps):
        date = start_date + datetime.timedelta(days=d)
        for _ in range(cap):
            if pos >= len(ordered):
                break
            out.append((date, ordered[pos]))
            pos += 1
    return out


def classify_buckets(anchors, smart_slot=None, cheap_slot=None) -> dict[str, str]:
    """Map distinct anchors -> cadence bucket.

    Only unique anchors are sent to the model (few even for thousands of links),
    so we use the *smart* model for them and fall back to the *cheap* model only
    for anchors it failed to label; anything still unlabeled is classified
    deterministically. Returns ``{anchor: bucket}``.
    """
    distinct = sorted(set(anchors))
    types: dict[str, str] = {}
    if smart_slot:
        types.update(anchortypes.llm_classify(distinct, smart_slot[0], smart_slot[1]))
    missing = [a for a in distinct if a not in types]
    if missing and cheap_slot:
        types.update(anchortypes.llm_classify(missing, cheap_slot[0], cheap_slot[1]))
    for a in distinct:  # deterministic fill for anything still unlabeled
        types.setdefault(a, anchortypes.classify_one(a))
    return {a: bucket_of(code) for a, code in types.items()}


def classify_with_model(anchors: list[str], key: str, model: str) -> dict[str, str]:
    """Back-compat single-model classification (distinct anchors -> buckets)."""
    return classify_buckets(anchors, smart_slot=(key, model))


def build_scheduled_workbook(header: list[str], placements: list[tuple[datetime.date, dict]],
                             date_fmt: str = "%d.%m.%Y", sheet_title: str = "Планнинг по датам") -> bytes:
    """Write the re-scheduled plan. Puts dates into the ``Sprint`` column
    (renamed to ``Date``), or into an existing ``Date`` column, or a new leading
    ``Date`` column when the source has neither."""
    if "Sprint" in header:
        date_i = header.index("Sprint")
        out_header = ["Date" if h == "Sprint" else h for h in header]
    elif "Date" in header:
        date_i = header.index("Date")
        out_header = header[:]
    else:
        date_i = None
        out_header = ["Date"] + header

    def make_row(date, link):
        row = link["row"][:]
        if date_i is not None:
            while len(row) <= date_i:
                row.append("")
            row[date_i] = date.strftime(date_fmt)
            return row
        return [date.strftime(date_fmt)] + row

    widths = [len(c) for c in out_header]
    for date, link in placements:
        for i, v in enumerate(make_row(date, link)):
            if i < len(widths):
                widths[i] = max(widths[i], len(str(v)))

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=re.sub(r"[\[\]:*?/\\]", "-", (sheet_title or "План"))[:31] or "План")
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(w + 4, 70)
    ws.freeze_panes = "A2"
    head = []
    for col in out_header:
        cell = WriteOnlyCell(ws, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        head.append(cell)
    ws.append(head)
    for date, link in placements:
        ws.append(make_row(date, link))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def summarize(links: list[dict], days: int) -> dict:
    """Quick stats for the UI / logs."""
    cats = defaultdict(int)
    for l in links:
        cats[l["category"]] += 1
    total = len(links)
    return {
        "total": total,
        "per_day": round(total / days, 1) if days else 0,
        "anchorless": cats[ANCHORLESS],
        "branded": cats[BRANDED],
        "commercial": cats[COMMERCIAL],
    }
