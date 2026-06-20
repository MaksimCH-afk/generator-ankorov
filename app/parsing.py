"""Parsing of uploaded frequency / project files (Excel or CSV)."""
from __future__ import annotations

import csv
import io
import re

from openpyxl import load_workbook

# Column-name synonyms used when importing full projects from one workbook.
KEYWORD_SYNS = ("keyword", "ключ", "запрос", "phrase", "key word", "keywords", "фраза", "анкор")
VOLUME_SYNS = ("volume", "vol", "частотн", "частот", "freq", "wordstat", "ws", "показ", "трафик")
# KD / Keyword Difficulty and similar are intentionally ignored.

_DOMAIN_RE = re.compile(r"^(https?://)?([a-z0-9-]+\.)+[a-z]{2,}(/.*)?$", re.I)


def _rows_from_xlsx(content: bytes) -> list[list[str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    return _normalize_rows(ws)


def _normalize_rows(ws) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c).strip() for c in row])
    return rows


def _rows_from_csv(content: bytes) -> list[list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    # Sniff delimiter (comma / semicolon / tab) from the first non-empty line.
    sample = next((ln for ln in text.splitlines() if ln.strip()), "")
    delimiter = ","
    for cand in [";", "\t", ","]:
        if cand in sample:
            delimiter = cand
            break
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [[c.strip() for c in row] for row in reader]


def read_table(filename: str, content: bytes) -> list[list[str]]:
    """Read an uploaded file into a list of string rows (first sheet for Excel)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return _rows_from_xlsx(content)
    return _rows_from_csv(content)


def _looks_like_header(cells: list[str]) -> bool:
    joined = " ".join(cells).lower()
    return any(w in joined for w in ("keyword", "ключ", "frequency", "freq", "частот", "volume", "объ", "vol"))


def rows_to_pairs(rows: list[list[str]]) -> list[tuple[str, float]]:
    """Turn raw string rows into ``[(keyword, frequency), ...]``.

    The first column is the keyword, the second (if any) the frequency/volume.
    A header row (if detected) is skipped. Rows without a keyword are ignored;
    a missing or non-numeric frequency defaults to 0.
    """
    out: list[tuple[str, float]] = []
    for i, cells in enumerate(rows):
        if not cells or not any(cells):
            continue
        if i == 0 and _looks_like_header(cells):
            continue
        keyword = cells[0].strip()
        if not keyword:
            continue
        freq = 0.0
        if len(cells) > 1:
            raw = cells[1].replace(",", ".").replace(" ", "")
            try:
                freq = float(raw)
            except ValueError:
                freq = 0.0
        out.append((keyword, freq))
    return out


def parse_frequency(filename: str, content: bytes) -> list[tuple[str, float]]:
    """Parse a single ``keyword | frequency`` table (first sheet for Excel)."""
    return rows_to_pairs(read_table(filename, content))


def parse_workbook_sheets(content: bytes) -> dict[str, list[tuple[str, float]]]:
    """Parse every sheet of an Excel workbook (simple first-two-columns mode)."""
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    result: dict[str, list[tuple[str, float]]] = {}
    for sheet_name in wb.sheetnames:
        result[sheet_name] = rows_to_pairs(_normalize_rows(wb[sheet_name]))
    return result


# --------------------------------------------------------------------------- #
# Smart project import (one workbook -> several projects)
# --------------------------------------------------------------------------- #
def _looks_like_domain(value: str) -> bool:
    v = value.strip()
    if not v or " " in v:
        return False
    return bool(_DOMAIN_RE.match(v))


def normalize_domain(value: str) -> str:
    """Normalise a domain/URL to ``https://host/...``.

    Forces https, drops a leading ``www.``, and ensures a trailing slash on the
    root form — so ``http://www.site.com`` and ``site.com/`` collapse to the same
    canonical URL (avoids duplicate projects).
    """
    v = value.strip()
    if not re.match(r"^https?://", v, re.I):
        v = "https://" + v
    v = re.sub(r"^http://", "https://", v, flags=re.I)
    v = re.sub(r"^https://www\.", "https://", v, flags=re.I)
    # Ensure the root form ends with a single slash.
    if "/" not in v.split("://", 1)[1]:
        v += "/"
    return v


def _find_header(rows: list[list[str]]) -> tuple[int | None, int | None, int | None]:
    """Locate the header row and the keyword / volume column indexes."""
    for idx, cells in enumerate(rows[:6]):
        low = [c.lower() for c in cells]
        kw_col = next((i for i, c in enumerate(low) if any(s in c for s in KEYWORD_SYNS)), None)
        if kw_col is None:
            continue
        vol_col = next((i for i, c in enumerate(low) if any(s in c for s in VOLUME_SYNS)), None)
        return idx, kw_col, vol_col
    return None, None, None


def _find_domain(rows: list[list[str]]) -> str | None:
    """Scan all cells for the first value that looks like a domain/URL."""
    for cells in rows:
        for c in cells:
            if _looks_like_domain(c):
                return normalize_domain(c)
    return None


def _to_float(raw: str) -> float:
    try:
        return float(raw.replace(",", ".").replace(" ", ""))
    except ValueError:
        return 0.0


def parse_project_sheet(rows: list[list[str]]) -> tuple[str | None, list[tuple[str, float]]]:
    """Parse one sheet into ``(domain, [(keyword, frequency), ...])``.

    Handles synonymous column names, columns in any position and a domain that
    can live in a header cell, a dedicated column or repeated in every row.
    Returns ``(None, [])`` for empty / summary / dashboard sheets.
    """
    rows = [r for r in rows if any(r)]
    if not rows:
        return None, []
    header_idx, kw_col, vol_col = _find_header(rows)
    domain = _find_domain(rows)
    if header_idx is None or kw_col is None:
        return domain, []  # no keyword column -> not a project sheet

    pairs: list[tuple[str, float]] = []
    for cells in rows[header_idx + 1:]:
        if kw_col >= len(cells):
            continue
        keyword = cells[kw_col].strip()
        if not keyword or _looks_like_domain(keyword):
            continue
        freq = _to_float(cells[vol_col]) if (vol_col is not None and vol_col < len(cells)) else 0.0
        pairs.append((keyword, freq))
    return domain, pairs


def parse_project_sheets(content: bytes) -> list[dict]:
    """Import projects from a workbook: one sheet = one project.

    Returns a list of ``{"name": sheet_name, "domain": str|None, "pairs": [...]}``.
    Empty and summary sheets yield empty ``pairs`` and are skipped by the caller.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        domain, pairs = parse_project_sheet(_normalize_rows(wb[sheet_name]))
        out.append({"name": sheet_name, "domain": domain, "pairs": pairs})
    return out


def parse_project_table(filename: str, content: bytes) -> dict:
    """Same as :func:`parse_project_sheet` but for a single CSV/Excel table."""
    rows = read_table(filename, content)
    domain, pairs = parse_project_sheet(rows)
    return {"name": filename, "domain": domain, "pairs": pairs}
