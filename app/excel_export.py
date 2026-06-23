"""Excel (and batch ZIP) export."""
from __future__ import annotations

import io
import re
import zipfile

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .generator import GeneratedRow, domain_of

# Output columns of the final TЗ (see reference file). One row = one link.
BASE_COLUMNS = [
    "Sprint",
    "SEO Specialist",
    "Project",
    "Project Url",
    "URL Type",
    "Link Type",
    "Anchor Type",
    "Anchor",
    "Keyword",
]
LANG_COLUMN = "Article Language"

# Match the reference file: black header row, white bold text, black body text.
HEADER_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BODY_FONT = Font(color="000000")

INTERNAL_SHEET = "Внутренние страницы"


def _line(row: GeneratedRow, sprint: str, seo: str, url_type: str,
          include_language: bool, language: str, keyword: str) -> list:
    """Build one output line. Link Type / Anchor Type stay empty; Keyword holds
    the project's most-used keyword (same value on every row)."""
    line = [
        sprint,
        seo,
        domain_of(row.url),   # Project: bare domain (site.com)
        row.url,              # Project Url: the page we link to
        url_type,
        "",                   # Link Type — empty
        "",                   # Anchor Type — empty
        row.anchor,           # Anchor — as computed
        keyword,              # Keyword — project's top keyword, on every row
    ]
    if include_language:
        line.append(language)
    return line


def top_keyword(sheets: dict[str, list[GeneratedRow]]) -> str:
    """The most-used keyword anchor across a project (highest link count).

    Bare-URL/domain (anchorless) and internal-page rows don't count — only real
    frequency keywords. Empty for a fully anchorless campaign.
    """
    best_count = 0
    best = ""
    for rows in sheets.values():
        for row in rows:
            if getattr(row, "is_keyword", False) and row.link_qty > best_count:
                best_count = row.link_qty
                best = row.keyword
    return best


def _write_sheet(wb, ws, rows: list[GeneratedRow], columns: list[str], sprint: str, seo: str,
                 url_type: str, include_language: bool, language: str, grouped: bool,
                 keyword: str) -> None:
    # Column widths first (write-only mode wants dimensions before rows).
    widths = [len(c) for c in columns]
    for row in rows:
        line = _line(row, sprint, seo, url_type, include_language, language, keyword)
        values = ([str(row.link_qty)] + line) if grouped else line
        for i, value in enumerate(values):
            widths[i] = max(widths[i], len(str(value)))
    for i, width in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = min(width + 4, 70)
    ws.freeze_panes = "A2"

    # Styled header row (black fill, white bold) via write-only cells.
    header = []
    for col in columns:
        cell = WriteOnlyCell(ws, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        header.append(cell)
    ws.append(header)

    # Body — streamed row by row to keep memory flat for large volumes.
    for row in rows:
        line = _line(row, sprint, seo, url_type, include_language, language, keyword)
        if grouped:
            if row.link_qty > 0:
                ws.append([row.link_qty] + line)
        else:
            for _ in range(max(0, row.link_qty)):
                ws.append(line)


def build_workbook(sheets: dict[str, list[GeneratedRow]], *, sprint: str = "",
                   seo_specialist: str = "", language: str = "",
                   include_language: bool | None = None, grouped: bool = False) -> bytes:
    """Build one .xlsx file. ``sheets`` maps sheet name -> rows.

    Uses openpyxl write-only (streaming) mode so even tens of thousands of
    expanded rows stay light on memory. By default each link is its own row;
    with ``grouped=True`` each anchor is a single row plus a leading
    ``Link Q-ty`` column. ``Article Language`` is appended when a language is set.
    """
    if include_language is None:
        include_language = bool((language or "").strip())
    columns = (["Link Q-ty"] if grouped else []) + BASE_COLUMNS + ([LANG_COLUMN] if include_language else [])

    keyword = top_keyword(sheets)
    wb = Workbook(write_only=True)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=_safe_sheet_name(name))
        url_type = "Inner Page" if name == INTERNAL_SHEET else "Main Page"
        _write_sheet(wb, ws, rows, columns, sprint, seo_specialist, url_type,
                     include_language, language, grouped, keyword)
    if not wb.sheetnames:  # never leave an empty workbook
        wb.create_sheet(title="Empty")
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no ``[]:*?/\\``."""
    cleaned = re.sub(r"[\[\]:\*\?/\\]", "-", name)
    return cleaned[:31] or "Sheet"


def safe_filename(url: str) -> str:
    """Turn a URL into a filesystem-safe .xlsx base name."""
    name = re.sub(r"^https?://", "", url).strip("/")
    name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
    return (name or "project") + ".xlsx"


def build_zip(files: dict[str, bytes]) -> bytes:
    """Bundle ``filename -> bytes`` into a ZIP archive."""
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, content in files.items():
            zf.writestr(filename, content)
    return buffer.getvalue()
